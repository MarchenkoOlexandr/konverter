# -*- coding: utf-8 -*-
#створення программи для конвертації кошторисів ф1 формату xlsx у текстовий формат для введення у программу АВК5
#команда для створення ехе файлу pyinstaller --onefile konverter.py
#підключаємо модуль для вибору файлів
import easygui
#вибираємо файл з початковим кошторисом у форматі ексель 
input_file = easygui.fileopenbox('Оберіть xlsx файл кошторису для обробки','', '*.xlsx')

#підключаємо модуль для читання ексель файлів
from openpyxl import load_workbook

#підключаємо модуль для роботи з строками
import re
import string
#завантажуємо книгу ексель з файлу
wb = load_workbook(input_file)
#обираємо місце для зберігання кінцевого файлу формату тхт
output_file = open (easygui.filesavebox( msg= 'Оберіть місце збереження файлу з результатами обробки', default= 'koshtorys.txt'), 'w+')
# отримуємо активний аркуш
wb.active = 0
sheet = wb.active
#визначаємо максимальну кількість строк
rows = sheet.max_row
#cols = sheet.max_column
for i in range (1, rows + 1):
	number = (sheet ['A'+str(i)].value)
	if number != None :
		obgryntyvannja = str(sheet ['B'+str(i)].value)
		vymir = str(sheet ['D'+str(i)].value)
		objem = str(sheet ['E'+str(i)].value)
		regex = re.compile('\n')
		objem = regex.sub('', objem)
		regex = re.compile('\n')
		obgryntyvannja = regex.sub('', obgryntyvannja)
		regex = re.compile('& ')
		obgryntyvannja = regex.sub('', obgryntyvannja)		
		for n in range (1, 20):
			variant = 'варіант '+str(n)
			obgryntyvannja = obgryntyvannja.replace(variant, '')
		if vymir == '100м2' or vymir == '100м3' or vymir == '100м' or vymir == '100т' or vymir == '100 м2' or vymir == '100 м3' or vymir == '100 м' or vymir == '100 т':
			objem = float(objem)*100
			objem = str(round(objem, 3))	
		if vymir == '1000м2' or vymir == '1000м3' or vymir == '1000м' or vymir == '1000т' or vymir == '1000 м2' or vymir == '1000 м3' or vymir == '1000 м' or vymir == '1000 т':
			objem = float(objem)*1000
			objem = str(round(objem, 3))
		if vymir == '10м2' or vymir == '10м3' or vymir == '10м' or vymir == '10т' or vymir == '10 м2' or vymir == '10 м3' or vymir == '10 м' or vymir == '10 т':
			objem = float(objem)*10
			objem = str(round(objem, 3))
		objem = objem.replace('.', ',')
		print (number, obgryntyvannja, vymir, objem)
		output_file.write(':П`'+str(obgryntyvannja)+'`'+str(objem)+'*\n')
output_file.close()
